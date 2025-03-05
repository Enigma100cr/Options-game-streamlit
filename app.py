import streamlit as st
import pandas as pd
from datetime import datetime
import plotly.express as px
import base64
import sqlite3
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import calendar
from fpdf import FPDF
from PIL import Image as PILImage
import io

# Database configuration
DATABASE_URI = 'sqlite:///trading_data.db'
conn = sqlite3.connect("trading_data.db", check_same_thread=False)
c = conn.cursor()

# Create tables
c.execute("""
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT UNIQUE,
    password TEXT,
    is_owner BOOLEAN DEFAULT FALSE
)
""")

c.execute("""
CREATE TABLE IF NOT EXISTS trades (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER,
    date TEXT,
    symbol TEXT,
    trade_type TEXT,
    entry_price REAL,
    exit_price REAL,
    stop_loss REAL,
    target REAL,
    qty INTEGER,
    status TEXT,
    setup_type TEXT,
    market_condition TEXT,
    psychology TEXT,
    notes TEXT,
    entry_screenshot BLOB,
    exit_screenshot BLOB,
    net_pnl REAL,
    FOREIGN KEY(user_id) REFERENCES users(id)
""")
conn.commit()

# Utility functions
def calculate_position_size(capital, risk_percent, entry, stop_loss, position_type):
    risk_amount = capital * (risk_percent / 100)
    risk_per_share = abs(entry - stop_loss)
    return round(risk_amount / risk_per_share) if risk_per_share != 0 else 0

def get_image_base64(image_file):
    if image_file is not None:
        return base64.b64encode(image_file.read()).decode()
    return None

def save_to_excel(trades_df, period, start_date, end_date, user_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Journal"
    headers = ["Date", "Symbol", "Type", "Entry", "Exit", "Qty", "Status", "Notes"]
    ws.append(headers)
    
    # Create a directory to store screenshots
    screenshot_dir = f"screenshots_user_{user_id}_{period}_{start_date}_{end_date}"
    os.makedirs(screenshot_dir, exist_ok=True)
    
    for index, row in trades_df.iterrows():
        # Save screenshots as images
        if row['entry_screenshot']:
            entry_screenshot_path = os.path.join(screenshot_dir, f"entry_{index}.png")
            with open(entry_screenshot_path, "wb") as f:
                f.write(base64.b64decode(row['entry_screenshot']))
        
        if row['exit_screenshot']:
            exit_screenshot_path = os.path.join(screenshot_dir, f"exit_{index}.png")
            with open(exit_screenshot_path, "wb") as f:
                f.write(base64.b64decode(row['exit_screenshot']))
        
        # Add trade data to the Excel file
        ws.append([
            row['date'], row['symbol'], row['trade_type'],
            row['entry_price'], row['exit_price'], row['qty'],
            row['status'], row['notes']
        ])
    
    excel_file = f"trade_journal_user_{user_id}_{period}_{start_date}_{end_date}.xlsx"
    wb.save(excel_file)
    return excel_file, screenshot_dir

def save_to_pdf(trades_df, user_id):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    
    # Add a title
    pdf.cell(200, 10, txt="Trade Journal", ln=True, align="C")
    pdf.ln(10)
    
    # Add trade details
    for _, row in trades_df.iterrows():
        pdf.cell(200, 10, txt=f"Date: {row['date']}", ln=True)
        pdf.cell(200, 10, txt=f"Symbol: {row['symbol']}", ln=True)
        pdf.cell(200, 10, txt=f"Type: {row['trade_type']}", ln=True)
        pdf.cell(200, 10, txt=f"Entry: {row['entry_price']}", ln=True)
        pdf.cell(200, 10, txt=f"Exit: {row['exit_price']}", ln=True)
        pdf.cell(200, 10, txt=f"Qty: {row['qty']}", ln=True)
        pdf.cell(200, 10, txt=f"Status: {row['status']}", ln=True)
        pdf.cell(200, 10, txt=f"Notes: {row['notes']}", ln=True)
        pdf.ln(10)
        
        # Add entry screenshot
        if row['entry_screenshot']:
            entry_screenshot = io.BytesIO(base64.b64decode(row['entry_screenshot']))
            img = PILImage.open(entry_screenshot)
            img_path = f"entry_{row['id']}.png"
            img.save(img_path)
            pdf.image(img_path, x=10, y=pdf.get_y(), w=100)
            pdf.ln(80)
            os.remove(img_path)
        
        # Add exit screenshot
        if row['exit_screenshot']:
            exit_screenshot = io.BytesIO(base64.b64decode(row['exit_screenshot']))
            img = PILImage.open(exit_screenshot)
            img_path = f"exit_{row['id']}.png"
            img.save(img_path)
            pdf.image(img_path, x=10, y=pdf.get_y(), w=100)
            pdf.ln(80)
            os.remove(img_path)
    
    pdf_file = f"trade_journal_user_{user_id}.pdf"
    pdf.output(pdf_file)
    return pdf_file

# Function to generate calendar view
def generate_calendar_view(trades_df, year, month):
    cal = calendar.monthcalendar(year, month)
    month_name = calendar.month_name[month]
    st.subheader(f"📅 {month_name} {year} - Trade Calendar")
    
    # Create a DataFrame for the calendar
    calendar_df = pd.DataFrame(cal, columns=["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"])
    calendar_df = calendar_df.replace(0, "")  # Replace 0s with empty strings
    
    # Add trade data to the calendar
    for index, row in trades_df.iterrows():
        trade_date = datetime.strptime(row['date'], "%Y-%m-%d").date()
        if trade_date.year == year and trade_date.month == month:
            day = trade_date.day
            for i, week in enumerate(cal):
                if day in week:
                    row_idx = i
                    col_idx = week.index(day)
                    pnl = row['net_pnl']
                    color = "green" if pnl > 0 else "red"
                    calendar_df.iloc[row_idx, col_idx] = f"<span style='color: {color};'>{day}</span>"
    
    # Display the calendar
    st.markdown(calendar_df.to_html(escape=False), unsafe_allow_html=True)

# Main app
st.set_page_config(page_title="Professional Trading Journal", layout="wide")

# Login system
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.user_id = None
    st.session_state.is_owner = False

def login(username, password):
    c.execute("SELECT id, password, is_owner FROM users WHERE username = ?", (username,))
    user = c.fetchone()
    if user and user[1] == password:
        st.session_state.logged_in = True
        st.session_state.user_id = user[0]
        st.session_state.is_owner = user[2]
        return True
    return False

def logout():
    st.session_state.logged_in = False
    st.session_state.user_id = None
    st.session_state.is_owner = False

# Login page
if not st.session_state.logged_in:
    st.title("Login to Trading Journal")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    if st.button("Login"):
        if login(username, password):
            st.success("Logged in successfully!")
        else:
            st.error("Invalid username or password.")
    st.markdown("---")
    st.write("Don't have an account? Contact the owner to create one.")
else:
    # Logout button
    if st.button("Logout"):
        logout()
        st.rerun()

    # Main app
    if st.session_state.logged_in:
        st.title(f"📈 Trading Journal - User {st.session_state.user_id}")
        st.markdown("---")

        # Tabs setup
        tabs = st.tabs(["📝 Trade Journal", "🧮 Position Calculator", "📊 Analytics", "⚙️ Settings"])

        with tabs[0]:  # Trade Journal
            col1, col2 = st.columns([1, 2])
            with col1:
                st.subheader("🔍 Filter Trades")
                start_date = st.date_input("Start Date", datetime.today())
                end_date = st.date_input("End Date", datetime.today())
                selected_symbol = st.text_input("Filter by Symbol")
            
            with col2:
                st.subheader("📜 Trade History")
                query = """
                    SELECT * FROM trades 
                    WHERE date BETWEEN ? AND ?
                    AND symbol LIKE ?
                    AND user_id = ?
                """
                trades_df = pd.read_sql(query, conn, params=(
                    start_date.strftime("%Y-%m-%d"),
                    end_date.strftime("%Y-%m-%d"),
                    f"%{selected_symbol}%",
                    st.session_state.user_id
                ))
                
                if not trades_df.empty:
                    # Download Options
                    st.subheader("📥 Download Options")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.write("**Day-wise Download**")
                        day_start = st.date_input("Start Date (Day-wise)", datetime.today(), key="day_start")
                        day_end = st.date_input("End Date (Day-wise)", datetime.today(), key="day_end")
                        if st.button("Download Day-wise"):
                            excel_file, screenshot_dir = save_to_excel(trades_df, "daywise", day_start, day_end, st.session_state.user_id)
                            with open(excel_file, "rb") as f:
                                st.download_button(
                                    label="⬇️ Download Day-wise",
                                    data=f,
                                    file_name="trade_journal_daywise.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                    with col2:
                        st.write("**Month-wise Download**")
                        month_start = st.date_input("Start Date (Month-wise)", datetime.today(), key="month_start")
                        month_end = st.date_input("End Date (Month-wise)", datetime.today(), key="month_end")
                        if st.button("Download Month-wise"):
                            excel_file, screenshot_dir = save_to_excel(trades_df, "monthwise", month_start, month_end, st.session_state.user_id)
                            with open(excel_file, "rb") as f:
                                st.download_button(
                                    label="⬇️ Download Month-wise",
                                    data=f,
                                    file_name="trade_journal_monthwise.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                    with col3:
                        st.write("**Year-wise Download**")
                        year_start = st.date_input("Start Date (Year-wise)", datetime.today(), key="year_start")
                        year_end = st.date_input("End Date (Year-wise)", datetime.today(), key="year_end")
                        if st.button("Download Year-wise"):
                            excel_file, screenshot_dir = save_to_excel(trades_df, "yearwise", year_start, year_end, st.session_state.user_id)
                            with open(excel_file, "rb") as f:
                                st.download_button(
                                    label="⬇️ Download Year-wise",
                                    data=f,
                                    file_name="trade_journal_yearwise.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                    
                    # PDF Export
                    if st.button("Download PDF"):
                        pdf_file = save_to_pdf(trades_df, st.session_state.user_id)
                        with open(pdf_file, "rb") as f:
                            st.download_button(
                                label="⬇️ Download PDF",
                                data=f,
                                file_name="trade_journal.pdf",
                                mime="application/pdf"
                            )
                    
                    # Calendar View
                    st.subheader("📅 Trade Calendar")
                    selected_year = st.selectbox("Select Year", [2024, 2025])
                    selected_month = st.selectbox("Select Month", list(range(1, 13)), format_func=lambda x: calendar.month_name[x])
                    generate_calendar_view(trades_df, selected_year, selected_month)
                    
                    # Trade History Table
                    for _, trade in trades_df.iterrows():
                        with st.expander(f"{trade['symbol']} - {trade['date']} - {trade['status']}"):
                            cols = st.columns([3,1])
                            with cols[0]:
                                st.write(f"**Entry:** ₹{trade['entry_price']} | **Exit:** ₹{trade['exit_price']}")
                                st.write(f"**Qty:** {trade['qty']}")
                                st.write(f"**Net P&L:** ₹{trade['net_pnl']:,.2f}")
                                st.write(f"**Notes:** {trade['notes']}")
                            with cols[1]:
                                if trade['entry_screenshot']:
                                    st.image(base64.b64decode(trade['entry_screenshot']), use_container_width=True)
                                if trade['exit_screenshot']:
                                    st.image(base64.b64decode(trade['exit_screenshot']), use_container_width=True)
                                if st.button("✏️ Edit", key=f"edit_{trade['id']}"):
                                    st.session_state.edit_trade = trade
                                if st.button("🗑️ Delete", key=f"delete_{trade['id']}"):
                                    c.execute("DELETE FROM trades WHERE id=?", (trade['id'],))
                                    conn.commit()
                                    st.rerun()
                else:
                    st.info("No trades found for selected filters")

        with tabs[1]:  # Position Calculator
            st.subheader("📐 Position Size Calculator")
            
            # Long Positions
            st.markdown("### Long Positions")
            col1, col2 = st.columns(2)
            with col1:
                long_entry = st.number_input("Entry Price (₹)", value=100.0, key="long_entry")
                long_stop = st.number_input("Stop Price (₹)", value=80.0, key="long_stop")
                long_target = st.number_input("Target Price (₹)", value=150.0, key="long_target")
            with col2:
                long_risk = st.number_input("Percent Risk (%)", value=2.0, key="long_risk")
                long_capital = st.number_input("Account Size (₹)", value=100000.0, key="long_capital")
            
            if st.button("Calculate Long"):
                position_size = calculate_position_size(long_capital, long_risk, long_entry, long_stop, "Long")
                cost_of_position = position_size * long_entry
                trade_risk = long_capital * (long_risk / 100)
                reward_risk = (long_target - long_entry) / (long_entry - long_stop)
                
                st.markdown("### Outputs")
                st.write(f"**Quantity to Buy:** {position_size}")
                st.write(f"**Cost of Position:** ₹{cost_of_position:,.2f}")
                st.write(f"**Trade Risk:** ₹{trade_risk:,.2f}")
                st.write(f"**Reward/Risk Ratio:** {reward_risk:.2f}")

            st.markdown("---")

            # Short Positions
            st.markdown("### Short Positions")
            col3, col4 = st.columns(2)
            with col3:
                short_entry = st.number_input("Entry Price (₹)", value=200.0, key="short_entry")
                short_stop = st.number_input("Stop Price (₹)", value=220.0, key="short_stop")
                short_target = st.number_input("Target Price (₹)", value=140.0, key="short_target")
            with col4:
                short_risk = st.number_input("Percent Risk (%)", value=2.0, key="short_risk")
                short_capital = st.number_input("Account Size (₹)", value=50000.0, key="short_capital")
            
            if st.button("Calculate Short"):
                position_size = calculate_position_size(short_capital, short_risk, short_entry, short_stop, "Short")
                cost_of_position = position_size * short_entry
                trade_risk = short_capital * (short_risk / 100)
                reward_risk = (short_entry - short_target) / (short_stop - short_entry)
                
                st.markdown("### Outputs")
                st.write(f"**Quantity to Sell:** {position_size}")
                st.write(f"**Cost of Position:** ₹{cost_of_position:,.2f}")
                st.write(f"**Trade Risk:** ₹{trade_risk:,.2f}")
                st.write(f"**Reward/Risk Ratio:** {reward_risk:.2f}")

        with tabs[2]:  # Analytics
            st.subheader("📊 Performance Analytics")
            trades_df = pd.read_sql("SELECT * FROM trades WHERE user_id = ?", 
                                  conn, params=(st.session_state.user_id,))
            
            if not trades_df.empty:
                col1, col2, col3 = st.columns(3)
                with col1:
                    total_trades = len(trades_df)
                    st.metric("Total Trades", total_trades)
                with col2:
                    win_rate = (len(trades_df[trades_df['net_pnl'] > 0]) / total_trades) * 100
                    st.metric("Win Rate", f"{win_rate:.1f}%")
                with col3:
                    total_pnl = trades_df['net_pnl'].sum()
                    st.metric("Total P&L", f"₹{total_pnl:,.2f}")
                
                # Equity Curve
                st.plotly_chart(px.line(trades_df, x='date', y='net_pnl', title='Equity Curve'))
                
                # Win Rate vs. Loss Rate
                win_loss_df = trades_df.groupby(trades_df['net_pnl'] > 0).size().reset_index(name='count')
                win_loss_df['result'] = win_loss_df['net_pnl'].apply(lambda x: 'Win' if x else 'Loss')
                st.plotly_chart(px.pie(win_loss_df, names='result', values='count', title='Win Rate vs. Loss Rate'))
                
                # P&L Distribution
                st.plotly_chart(px.histogram(trades_df, x='net_pnl', title='P&L Distribution'))
            else:
                st.info("No data available for analytics")

        with tabs[3]:  # Settings/New Trade
            st.subheader("➕ New Trade Entry")
            
            with st.form("new_trade_form"):
                col1, col2 = st.columns(2)
                with col1:
                    trade_date = st.date_input("Trade Date")
                    symbol = st.text_input("Symbol")
                    trade_type = st.selectbox("Type", ["Long", "Short"])
                    entry_price = st.number_input("Entry Price")
                    exit_price = st.number_input("Exit Price")
                    stop_loss = st.number_input("Stop Loss")
                    target_price = st.number_input("Target Price")
                    qty = st.number_input("Qty", min_value=1, value=1)  # Added Qty field
                
                with col2:
                    status = st.selectbox("Status", ["Open", "Closed"])
                    setup_type = st.selectbox("Setup Type", ["Breakout", "Reversal", "Trend"])
                    market_condition = st.selectbox("Market Condition", ["Bullish", "Bearish", "Sideways"])
                    psychology = st.selectbox("Psychology", ["Confident", "Fearful", "Revenge"])
                    entry_screenshot = st.file_uploader("Entry Screenshot", type=["png", "jpg", "jpeg"])
                    exit_screenshot = st.file_uploader("Exit Screenshot", type=["png", "jpg", "jpeg"])
                
                notes = st.text_area("Trade Notes")
                
                if st.form_submit_button("Save Trade"):
                    # Correct P&L calculation based on trade type
                    if trade_type == "Long":
                        pnl = (exit_price - entry_price) * qty  # Qty used for P&L calculation
                    else:  # Short
                        pnl = (entry_price - exit_price) * qty  # Qty used for P&L calculation
                    net_pnl = pnl  # No brokerage
                    
                    c.execute("""
                        INSERT INTO trades (
                            user_id, date, symbol, trade_type, entry_price, exit_price,
                            stop_loss, target, qty, status, setup_type,
                            market_condition, psychology, notes, entry_screenshot,
                            exit_screenshot, net_pnl
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        st.session_state.user_id,
                        trade_date.strftime("%Y-%m-%d"),
                        symbol,
                        trade_type,
                        entry_price,
                        exit_price,
                        stop_loss,
                        target_price,
                        qty,  # Added Qty field
                        status,
                        setup_type,
                        market_condition,
                        psychology,
                        notes,
                        get_image_base64(entry_screenshot),
                        get_image_base64(exit_screenshot),
                        net_pnl
                    ))
                    conn.commit()
                    st.success("Trade saved successfully!")

        # Edit Trade Modal
        if 'edit_trade' in st.session_state:
            trade = st.session_state.edit_trade
            with st.form("edit_trade_form"):
                st.subheader("Edit Trade")
                
                col1, col2 = st.columns(2)
                with col1:
                    new_entry = st.number_input("Entry Price", value=trade['entry_price'])
                    new_exit = st.number_input("Exit Price", value=trade['exit_price'])
                    new_stop = st.number_input("Stop Loss", value=trade['stop_loss'])
                    new_target = st.number_input("Target", value=trade['target'])
                    new_qty = st.number_input("Qty", value=trade['qty'], min_value=1)  # Added Qty field
                
                with col2:
                    new_status = st.selectbox("Status", ["Open", "Closed"], index=0 if trade['status'] == "Open" else 1)
                    new_notes = st.text_area("Notes", value=trade['notes'])
                
                if st.form_submit_button("Save Changes"):
                    c.execute("""
                        UPDATE trades SET
                        entry_price = ?,
                        exit_price = ?,
                        stop_loss = ?,
                        target = ?,
                        qty = ?,
                        status = ?,
                        notes = ?
                        WHERE id = ?
                    """, (
                        new_entry,
                        new_exit,
                        new_stop,
                        new_target,
                        new_qty,  # Added Qty field
                        new_status,
                        new_notes,
                        trade['id']
                    ))
                    conn.commit()
                    del st.session_state.edit_trade
                    st.rerun()

# Footer
st.markdown("---")
st.markdown("© 2023 Trading Journal App. All rights reserved.")
